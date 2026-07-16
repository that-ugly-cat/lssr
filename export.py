"""
Excel (.xlsx) export for the screening and assessment tabs, via openpyxl.

Screening: one row per record in the (non-removed) pool — bibliographic fields
plus the resolved screen-1 decision and a summary of every reviewer's vote.

Assessment: one row per screen-1-included record — bibliography, full-text status,
the resolved screen-2 decision + votes, and one column per extraction field
holding the authoritative value (final › latest user › model draft), with fields
hidden by an unmet show_if left blank.
"""
import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from models import (Extraction, Record, ScreenDecision, User, authoritative_values,
                    field_visible, workspace_extraction_fields)


def _as_list(js) -> list:
    try:
        v = json.loads(js) if js else []
        return v if isinstance(v, list) else []
    except (ValueError, TypeError):
        return []


def _fmt_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return "; ".join(str(x) for x in v if x not in (None, ""))
    return str(v)


def _user_names(db) -> dict:
    return {u.id: (u.name or u.email) for u in db.query(User).all()}


def _votes_by_record(db, stage: str, rec_ids: list) -> dict:
    out: dict = {}
    if not rec_ids:
        return out
    for v in (db.query(ScreenDecision)
                .filter(ScreenDecision.stage == stage,
                        ScreenDecision.record_id.in_(rec_ids)).all()):
        out.setdefault(v.record_id, []).append(v)
    return out


def _summarize_votes(rows, names: dict) -> str:
    def who(v):
        if v.reviewer_kind == "model":
            return "model"
        if v.reviewer_kind == "adjudicator":
            return "adjudicator"
        return names.get(v.reviewer_id, "user")
    return "; ".join(f"{who(v)}={v.decision}"
                     for v in sorted(rows, key=lambda r: (r.reviewer_kind, r.id)))


def _new_sheet(wb, title, headers):
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    return ws


def _bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BIB = ["Record ID", "Type", "Title", "Authors", "Year", "DOI", "URL", "Source", "Databases"]


def _bib_cells(r: Record) -> list:
    return [r.id, r.type, r.title, r.authors, r.year, r.doi, r.url, r.source,
            "; ".join(_as_list(r.source_dbs_json))]


def screening_xlsx(db, workspace) -> bytes:
    recs = (db.query(Record)
              .filter(Record.workspace_id == workspace.id, Record.is_removed == False)  # noqa: E712
              .order_by(Record.year.is_(None), Record.year.desc(), Record.id).all())
    names = _user_names(db)
    votes = _votes_by_record(db, "screen1", [r.id for r in recs])

    wb = Workbook()
    ws = _new_sheet(wb, "Screening",
                    _BIB + ["Language", "Abstract",
                            "Screen-1 decision", "Decided by", "Reason", "Reviewer votes"])
    for r in recs:
        ws.append(_bib_cells(r) + [
            r.language, r.abstract,
            r.screen1_decision, r.screen1_by, r.screen1_reason,
            _summarize_votes(votes.get(r.id, []), names),
        ])
    return _bytes(wb)


def assessment_xlsx(db, workspace) -> bytes:
    recs = (db.query(Record)
              .filter(Record.workspace_id == workspace.id, Record.is_removed == False,  # noqa: E712
                      Record.screen1_decision == "include")
              .order_by(Record.year.is_(None), Record.year.desc(), Record.id).all())
    fields = workspace_extraction_fields(db, workspace)
    names = _user_names(db)
    votes = _votes_by_record(db, "screen2", [r.id for r in recs])

    wb = Workbook()
    headers = (_BIB + ["Full-text status",
                       "Screen-2 decision", "Decided by", "Reason", "Reviewer votes"]
               + [f.label for f in fields])
    ws = _new_sheet(wb, "Assessment", headers)
    for r in recs:
        vals = authoritative_values(db, r)
        field_cells = [
            _fmt_value(vals.get(f.key)) if field_visible(f, vals) else ""
            for f in fields
        ]
        ws.append(_bib_cells(r) + [
            r.full_text_status,
            r.screen2_decision, r.screen2_by, r.screen2_reason,
            _summarize_votes(votes.get(r.id, []), names),
        ] + field_cells)
    return _bytes(wb)
